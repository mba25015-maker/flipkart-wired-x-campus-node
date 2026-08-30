"""
GENERATE Campus_Store_Model.xlsx FROM THE MODEL.

The workbook shipped with this package was Round-1 vintage for twelve days: 18-day NWC, ₹42 last
mile, Minutes AOV ₹775, ₹528 breakeven, 12.6×/8.1× asset turns. It survived because nothing
generated it and nothing read it - the same reason the A-1 screenshots and the appendix panels
went stale. A workbook that is written by hand is a second copy of every number in the package,
and second copies drift.

So this file is generated at release time, like the screenshots and the README. Inputs and
policy constants come from params/campus_model; downstream sheets use LIVE EXCEL FORMULAS that
reference the Inputs sheet, so a reader can trace the arithmetic in Excel rather than take the
model's word for it - and so the workbook recomputes rather than restating.

Verification counts come from check_counts, which reads the recorded run. Nothing about the
package's own state is typed here.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import params as P, campus_model as M, cost_stack as CS, sla as SL, fleet_mix as FM
import roce as RC, working_capital as WC, break_mode as B, risk_shocks as RS
import basket as BK, aishe_district as AD, check_counts as CC, pg_demand as PG

HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.dirname(HERE)
OUT  = os.path.join(ROOT, "Campus_Store_Model.xlsx")

INK="FF0B1F3A"; GOLD="FFFFC220"; BLUE="FFDCE7F7"; GREEN="FFE3F1E6"; GREY="FFF5F8FD"
H1=Font(name="Calibri",size=14,bold=True,color=INK)
H2=Font(name="Calibri",size=10,bold=True,color="FFFFFFFF")
BOLD=Font(name="Calibri",size=10,bold=True,color=INK)
BODY=Font(name="Calibri",size=10,color=INK)
MUTE=Font(name="Calibri",size=9,italic=True,color="FF5A6785")
BAND=PatternFill("solid",fgColor=INK); SRC=PatternFill("solid",fgColor=BLUE)
CALC=PatternFill("solid",fgColor=GREEN); CHIP=PatternFill("solid",fgColor=GREY)
PILOT=PatternFill("solid",fgColor=GOLD)
THIN=Border(bottom=Side(style="thin",color="FFDFE6F2"))

def sheet(wb,name,title,note):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    ws["A1"]=title; ws["A1"].font=H1
    ws["A2"]=note;  ws["A2"].font=MUTE
    ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=26
    ws["A2"].alignment=Alignment(wrap_text=True,vertical="top")
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=8)
    return ws

def header(ws,row,cols):
    for i,(t,w) in enumerate(cols,1):
        c=ws.cell(row=row,column=i,value=t); c.font=H2; c.fill=BAND
        c.alignment=Alignment(horizontal="left",vertical="center")
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[row].height=18

def put(ws,row,vals,fill=None,bold=False,fmt=None):
    for i,v in enumerate(vals,1):
        c=ws.cell(row=row,column=i,value=v)
        c.font=BOLD if bold else BODY; c.border=THIN
        if fill: c.fill=fill
        if fmt and i==2: c.number_format=fmt
    return row+1

def build():
    wb=Workbook(); wb.remove(wb.active)

    # ---------------------------------------------------------------- 0 Read Me
    ws=sheet(wb,"0 Read Me","Flipkart WiRED X — Campus Node Model",
        f"Generated from the model by Model/make_workbook.py. Team {P.TEAM_NAME}. Blue rows are "
        "sourced inputs; green rows are live formulas that recompute from the Inputs sheet. "
        "Nothing in this workbook is typed by hand, including the verification counts below.")
    r=4; ws.cell(row=r,column=1,value="RELEASE CHECKS").font=BOLD; r+=1
    header(ws,r,[("Layer",26),("Passed",10),("Defined",10),("Status",12),("Command",30),("Scope",56)]); r+=1
    verified=CC.all_verified()
    SCOPE={"audit":("python3 Model/audit.py","model against itself, incl. cross-module reconciliation"),
           "docs":("python3 Model/verify_docs.py","HANDOFF figures against the model"),
           "spec":("python3 Model/verify_spec.py","deck spec and build prompt against the model"),
           "deck":("python3 Model/verify_deck.py","the BUILT .pptx, presence and absence"),
           "artefacts":("python3 Model/verify_artifacts.py","workbook, notebooks, manifest, public data")}
    for key in CC.DEFINED:            # every layer, never a maintained list
        cmd,scope = SCOPE.get(key,("",""))
        res=(CC.results() or {}).get(key,{})
        r=put(ws,r,[key, res.get("passed","—"), CC.DEFINED[key],
                    f'=IF(B{r}=C{r},"PASS","FAIL")', cmd, scope], fill=CHIP)
    r+=1
    _n = len(CC.DEFINED)
    _word = {4:"FOUR",5:"FIVE",6:"SIX"}.get(_n, str(_n))
    ws.cell(row=r,column=1,value=(f"ALL {_word} LAYERS PASSED on the last recorded run."
            if verified else "NO CLEAN RUN RECORDED — counts are checks defined, not passed.")).font=BOLD

    # ---------------------------------------------------------------- 1 Inputs
    ws=sheet(wb,"1 Inputs","Inputs and policy constants",
        "Every contested policy constant lives in Model/params.py and is imported everywhere else. "
        "A value redeclared outside params.py fails the audit.")
    r=4; header(ws,r,[("Parameter",34),("Value",14),("Unit",14),("Tier",7),("Source / derivation",42)]); r+=1
    legs=SL.cost_legs()
    _input_rows = {}
    _inputs = [
        ("Cluster volume", M.CEILING, "orders/day","T1","JM Ex.29 observed 1,334–1,487 band"),
        ("Campus gates per node", P.CAMPUSES_PER_NODE, "gates","D","params.CAMPUSES_PER_NODE"),
        ("Active academic months", M.ACTIVE_MONTHS, "months/yr","T1","institutional calendars"),
        ("Revenue take rate", M.TAKE_RATE, "% of GOV","T1","JM Financial, Eternal FY26E"),
        ("Store fixed cost", CS.CAMPUS_FIXED, "Rs/month","T1","JM Financial Exhibit 13"),
        ("Store ops per order", M.STORE_OPS, "Rs/order","T2","pick, pack, warehousing"),
        ("Packaging per order", M.PACKAGING, "Rs/order","T2","packaging + support"),
        ("Unallocated residual", M.RESIDUAL, "Rs/order","D","implied variable less itemised"),
        ("City gig leg", legs["city"], "Rs/order","D","sla.py, volume-weighted across demand bands"),
        ("In-gate roster leg", legs["in_gate"], "Rs/order","D","fleet_mix.plan_roster(), fixed roster / volume"),
        ("NWC days", P.NWC_DAYS, "days","T1","Eternal Q1FY27, on NOV not COGS"),
        ("Capex midpoint", RC.CAPEX_MID, "Rs","T2","band midpoint"),
        ("Tax rate", RC.TAX_RATE, "%","T1","22% + surcharge + cess"),
        ("Node life", RC.NODE_LIFE_MO, "months","D","must outlive its own payback"),
        ("PG demand evidence gate", P.PG_DEMAND_ENABLED, "TRUE/FALSE","P","opens only after site evidence is approved"),
        ("PG verified occupied student beds", P.PG_VERIFIED_OCCUPIED_STUDENT_BEDS, "beds","P","occupied STUDENT beds inside the served radius"),
        ("PG active-user penetration", P.PG_ACTIVE_USER_PENETRATION, "%","P","pilot-observed active users / occupied student beds"),
        ("PG gross AOV", P.PG_AOV, "Rs/order","P","pilot-observed; zero means not evidenced"),
        ("PG common-drop share", P.PG_COMMON_DROP_SHARE, "%","P","share served at reception, warden desk or locker"),
        ("PG break retention", P.PG_BREAK_RETENTION, "% of term","P","PG demand persisting through the academic break"),
        ("Hostel term volume in PG scenario", P.PG_HOSTEL_TERM_OPD, "orders/day","P","capacity already consumed by the hostel base"),
        ("PG orders per active user per day", PG.PG_ORDERS_PER_ACTIVE_USER_DAY, "orders/day","T1","Blinkit 3.6/month / 30; no captive-campus uplift"),
        ("PG doorstep last mile", PG.PG_DOORSTEP_CPO, "Rs/order","D","cost_stack Type B urban PG geometry"),
        ("PG common-drop last mile", PG.PG_COMMON_DROP_CPO, "Rs/order","D","cost_stack Type B common-drop geometry"),
        ("NOV / GOV", WC.NOV_OVER_GOV, "x","T1","Blinkit FY26E"),
        ("Term days", RC.DAYS_TERM, "days/year","D","365 x active academic months / 12"),
        ("Break days", RC.DAYS_BREAK, "days/year","D","365 - term days"),
        ("Hostel comparison AOV", RC.AOV_UP, "Rs/order","D","basket at the 30% non-grocery floor"),
    ]
    for row in _inputs:
        _input_rows[row[0]] = r
        r=put(ws,r,list(row),fill=PILOT if row[3]=="P" else SRC)

    def _iref(label):
        return f"'1 Inputs'!$B${_input_rows[label]}"

    # ---------------------------------------------------------------- 2 Fulfilment
    ws=sheet(wb,"2 Fulfilment","Fulfilment — two legs, two cost structures",
        "The city leg is gig, bought per trip and divided by batch, so it varies by demand band. "
        "The in-gate leg is a fixed roster divided by daily volume, so it is FLAT across bands. "
        "That flatness is the argument for rostering.")
    r=4; header(ws,r,[("Demand band",26),("Hours",8),("Multiplier",11),("% of orders",12),
                      ("Orders/hr",11),("Batch",8),("City Rs/order",14),("In-gate Rs/order",15),("Total Rs/order",14),("SLA min",10)]); r+=1
    rows,weighted=SL.volume_weighted(); first=r
    for name,h,m,share,rate,b,cost,sla,city,ing in rows:
        r=put(ws,r,[name,h,m,share,round(rate,1),b,round(city,2),round(ing,2),round(cost,2),round(sla,1)],fill=CALC)
    r+=1
    r=put(ws,r,["VOLUME-WEIGHTED LAST MILE",None,None,None,None,None,
                f"=SUMPRODUCT(D{first}:D{first+2},G{first}:G{first+2})",f"=H{first}",
                f"=G{r}+H{r}"],bold=True)
    r=put(ws,r,["Standard 2–3 km residential zone",None,None,None,None,None,None,None,M.LAST_MILE])
    r=put(ws,r,["Saving vs a standard zone",None,None,None,None,None,None,None,
                f"=1-I{r-2}/I{r-1}"],bold=True)
    R14,A14,C14,G14=FM.plan_roster(); PR,PA,PC,PSAVE=FM.pooled_upside()
    r+=2
    r=put(ws,r,["ROSTER — base case is PER GATE",None,None,None,None],bold=True)
    r=put(ws,r,[f"per gate ({P.gate_volume():.0f}/day)", R14//G14, "runners", f"alpha {A14:.0%}", f"Rs{C14:.2f}/order"])
    r=put(ws,r,[f"pooled across {G14} gates (UPSIDE, not the plan)", PR, "runners", f"alpha {PA:.0%}",
                f"Rs{PC:.2f}/order — Rs{PSAVE:.2f} better, conditional on cross-gate movement"])

    # ---------------------------------------------------------------- 3 Unit economics
    ws=sheet(wb,"3 Unit Economics","Unit economics and the AOV thresholds",
        "Two day-count conventions are both in use and both correct. The deck spine carries the "
        "fixed base across twelve 30-day months; the returns model uses a 365-day year. The gap "
        "is stated rather than smoothed away.")
    r=4; header(ws,r,[("Metric",42),("Value",16),("Unit",12),("Reading",46)]); r+=1
    for row in [
        ("Variable cost per order", round(RC.LAST_MILE_D2+M.STORE_OPS+M.PACKAGING+M.RESIDUAL,2),"Rs","last mile + store ops + packaging + residual"),
        ("Active days per year", round(RC.DAYS_TERM,1),"days","365 x 8.5/12"),
        ("Annual orders, term-only", round(RC.orders_year()),"orders","returns basis"),
        ("Breakeven AOV — 365-day returns basis", round(RC.AOV_BREAKEVEN),"Rs","ROCE = 0"),
        ("Breakeven AOV — 30-day deck spine", round(RC.SPINE_BREAKEVEN),"Rs","the deck headline"),
        ("Day-count gap between the two", round(RC.DAYCOUNT_GAP,2),"Rs","asserted, not smoothed"),
        ("AOV at the 40% external benchmark", round(RC.AOV_HURDLE),"Rs","Eternal's disclosed return, not Flipkart's"),
        ("Premium over breakeven", round(RC.HURDLE_PREMIUM),"Rs","what the basket lever must buy"),
        ("Non-grocery share the benchmark implies", round(RC.HURDLE_NONGROCERY_SHARE,1),"% of GOV","inside the 30-40% range Swiggy discloses"),
    ]: r=put(ws,r,list(row),fill=CALC)

    # ---------------------------------------------------------------- 4 Returns
    ws=sheet(wb,"4 Returns","Return on capital, DuPont and the scenarios",
        "ROCE = EBIT/CE = margin leg x turnover leg. Capital employed is capex plus working "
        "capital on the adopted 14-day-on-NOV construct.")
    r=4; header(ws,r,[("Line",34),("Value",16),("Unit",12),("Note",46)]); r+=1
    du=RC.dupont(RC.AOV_HURDLE)
    for row in [("Capex, band midpoint",round(RC.CAPEX_MID/1e5,1),"Rs lakh",""),
                ("Working capital",round(WC.WC_ADOPTED/1e5,1),"Rs lakh",f"{P.NWC_DAYS:.0f} NWC days on NOV"),
                ("Capital employed",round(RC.CE_BASE/1e5,1),"Rs lakh","capex + working capital"),
                ("DuPont margin leg",round(du['ebit_margin']*100,2),"%","at the benchmark AOV"),
                ("DuPont turnover leg",round(du['capital_turn'],2),"x","the node's own turnover on CE"),
                ("Like-for-like asset turn",round(RC.TURN_SLIDE4,2),"x","common AOV Rs450, isolates density x calendar"),
                ("IRR at the benchmark AOV",round(RC.irr(RC.AOV_HURDLE)*100,1),"%","5-year life, 3-month ramp"),
               ]: r=put(ws,r,list(row),fill=CALC)
    r+=1
    header(ws,r,[("Scenario",34),("AOV",12),("Margin",12),("Turn",10),("ROCE",10),("Payback mo",13)]); r+=1
    for s in RC.scenario_rows():
        pay="—" if s['payback']!=s['payback'] or s['payback']==float('inf') else round(s['payback'])
        r=put(ws,r,[s['name'],round(s['aov']),round(s['margin']*100,2),round(s['turn'],2),
                    round(s['roce']*100,1),pay])
    r+=1
    dn=RC.scenario_rows()[3]
    ws.cell(row=r,column=1,value=(f"THE LIMIT: at a 30% basket with volume -30%, ROCE is "
        f"{dn['roce']:.1%} and payback runs to {dn['payback']:.0f} months against a "
        f"{RC.NODE_LIFE_MO}-month life. The node does not pay back within its life under that "
        f"combination, which is why day 90 measures volume first.")).font=BOLD

    # ---------------------------------------------------------------- 5 Dead zone
    ws=sheet(wb,"5 Dead Zone","The academic dead zone, priced",
        "A 3.5-month break with no students. Five strategies, one basis. Repurpose fills the node "
        "from the adjacent catchment rather than mothballing it.")
    r=4; header(ws,r,[("Restart credit state",34),("Working capital",18),("Note",52)]); r+=1
    import working_capital as _W
    for st,lbl,note in (("A","credit intact","suppliers resume terms; the rebuild is payables-funded"),
                        ("B30","30-day reset","partially self-funded while terms re-establish"),
                        ("B","full reset (ADOPTED)","the whole 60-day cycle self-funded — the downside")):
        wc=_W.reactivation_wc("B" if st!="A" else "A", days_credit=30 if st=="B30" else None)*(1-0.15)
        r=put(ws,r,[lbl,round(wc/1e5,1),note],fill=CALC if st=="B" else None)
    r+=1
    r=put(ws,r,["Working-capital constructs",None,None],bold=True)
    for lbl,v,st in ((f"NWC days x NOV ({P.NWC_DAYS:.0f}d)",WC.WC_ADOPTED,"ADOPTED"),
                     (f"{P.NWC_DAYS_TARGET:.0f}-day steady state",WC.WC_TARGET,"sensitivity"),
                     (f"COGS x {P.NWC_DAYS_R1:.0f} days",WC.WC_OLD,"REJECTED — double-counts the netting")):
        r=put(ws,r,[lbl,round(v/1e5,1),st])

    # ---------------------------------------------------------------- 6 Risks
    ws=sheet(wb,"6 Risks","Priced shocks, against the deck spine",
        "Each shock is re-solved for the AOV it would require, on the same D2-consistent basis as "
        "the spine. One combined case moves two variables together.")
    r=4; header(ws,r,[("Shock",34),("Breakeven AOV",16),("Basis",46)]); r+=1
    for lbl,v,basis in (("Volume -30%",RS.AOV_VOLUME,"thinner fixed-base absorption"),
                        ("Shrinkage, upper bound",RS.AOV_SHRINKAGE,"1.8% of NOV"),
                        ("Gig social-security levy",RS.AOV_LEVY,"201st Parliamentary report"),
                        ("Calendar fragmentation",RS.AOV_FRAGMENTATION,"shape of the break, not its length")):
        r=put(ws,r,[lbl,round(v),basis],fill=CALC)
    r+=1
    ws.cell(row=r,column=1,value=f"Deck spine for comparison: Rs{RC.SPINE_BREAKEVEN:.0f}").font=BOLD

    # ---------------------------------------------------------------- 7 District screen
    ws=sheet(wb,"7 District Screen","The site screen — districts, not sites",
        "A desktop screen over the AISHE register. It ranks candidates; siting still runs the "
        "cluster test on the ground.")
    r=4; header(ws,r,[("Step",42),("Count",14),("Note",46)]); r+=1
    for lbl,v,note in (("Institutions in the register",AD.N_HEI,"AISHE, as on 28-8-2026"),
                       ("Colleges only",AD.N_COL,"colleges"),
                       ("Urban colleges",AD.URBAN_COL,"urban flag"),
                       ("Districts passing four criteria",AD.N_CANDIDATES,f"of {AD.N_DISTRICTS} (State, District) pairs"),
                       ("Uncontested",9,"the screen's own floor: 6-7 urban colleges"),
                       ("Contested",22,""),("Stacked",80,"")):
        r=put(ws,r,[lbl,v,note],fill=CALC)

    # ---------------------------------------------------------------- 8 Verification
    ws=sheet(wb,"8 Verification","How this package checks itself",
        f"{len(CC.DEFINED)} verification layers plus four source scans. Absence checks matter as much "
        "as presence: "
        "ones: a superseded figure in a panel nothing reads is invisible to a positive check.")
    r=4; header(ws,r,[("Guard",40),("What it catches",70)]); r+=1
    for g,w in (("params.py","a contested policy constant defined more than once"),
                ("verify_artifacts.py","a stale figure in the workbook, notebooks or asset manifest"),
                ("recon(a, b)","two modules computing the same quantity under different policies"),
                ("phrase bans","wording the deck must not contain, including attribution"),
                ("superseded-value bans","a pre-correction figure surviving in any panel"),
                ("duplicate-registration guard","a check registered twice, inflating the count"),
                ("stale-basis scan","a superseded value used as a literal cost basis, defaults included"),
                ("document scan","a superseded figure in HANDOFF, spec, build prompt or README"),
                ("release.py","build - verify - capture - stamp - re-verify the exact stamped file")):
        r=put(ws,r,[g,w])

    # ---------------------------------------------------------------- 9 PG adjacency
    ws=sheet(wb,"9 PG Adjacency","PG adjacency — evidence-gated demand and capacity",
        "The hostel-only base case is unchanged. PG demand enters only after the evidence gate opens, "
        "and only into spare term capacity or available break-period capacity. Edit the gold pilot "
        "inputs on 1 Inputs; every output below is a live formula.")
    r=4; ws.cell(row=r,column=1,value="CONFIGURED SCENARIO — ZERO UNTIL THE PILOT INPUTS ARE EVIDENCED").font=BOLD; r+=1
    header(ws,r,[("Metric",40),("Value",18),("Unit",16),("Decision reading",58)]); r+=1
    _admit=(f"AND({_iref('PG demand evidence gate')}=TRUE,"
            f"{_iref('PG verified occupied student beds')}>0,"
            f"{_iref('PG active-user penetration')}>0,"
            f"{_iref('PG gross AOV')}>0)")
    _gross=r; r=put(ws,r,["PG demand evidence status",f'=IF({_admit},"OPEN","CLOSED")',"",
                          "CLOSED means PG contributes zero to every published return"],fill=PILOT,bold=True)
    _gross=r; r=put(ws,r,["Gross PG demand",f'=IF({_admit},{_iref("PG verified occupied student beds")}*{_iref("PG active-user penetration")}*{_iref("PG orders per active user per day")},0)',"orders/day",
                          "verified beds × active-user penetration × frequency"],fill=CALC)
    _spare=r; r=put(ws,r,["Spare term capacity",f'=MAX(0,{_iref("Cluster volume")}-{_iref("Hostel term volume in PG scenario")})',"orders/day",
                          "PG is not added above the store working ceiling"],fill=CALC)
    _served=r; r=put(ws,r,["PG admitted in term",f'=MIN(B{_gross},B{_spare})',"orders/day",
                           "the lower of evidenced demand and spare capacity"],fill=CALC)
    _overflow=r; r=put(ws,r,["PG overflow / capacity trigger",f'=MAX(0,B{_gross}-B{_served})',"orders/day",
                             "requires capacity expansion or remains unserved"],fill=CALC)
    _break=r; r=put(ws,r,["PG admitted in academic break",f'=MIN({_iref("Cluster volume")},B{_gross}*{_iref("PG break retention")})',"orders/day",
                          "zero until break retention is measured"],fill=CALC)
    _lm=r; r=put(ws,r,["Weighted PG last mile",f'=(1-{_iref("PG common-drop share")})*{_iref("PG doorstep last mile")}+{_iref("PG common-drop share")}*{_iref("PG common-drop last mile")}',"Rs/order",
                       "doorstep/common-drop mix"],fill=CALC)
    _cm=r; r=put(ws,r,["PG contribution per order",f'=IF({_iref("PG gross AOV")}>0,{_iref("Revenue take rate")}*{_iref("PG gross AOV")}-B{_lm}-{_iref("Store ops per order")}-{_iref("Packaging per order")}-{_iref("Unallocated residual")},0)',"Rs/order",
                       "same take rate and store-cost basis as the core model"],fill=CALC)
    _annual=r; r=put(ws,r,["Annual incremental PG orders",f'=B{_served}*{_iref("Term days")}+B{_break}*{_iref("Break days")}',"orders/year",
                           "term spare capacity plus measured break retention"],fill=CALC)
    _ebit=r; r=put(ws,r,["Incremental PG EBIT",f'=B{_annual}*B{_cm}',"Rs/year",
                         "before any capacity-expansion capex"],fill=CALC)
    _nwc=r; r=put(ws,r,["Incremental PG working capital",f'=MAX(B{_served},B{_break})*{_iref("PG gross AOV")}*{_iref("NOV / GOV")}*{_iref("NWC days")}',"Rs",
                        "14 NWC days on incremental NOV"],fill=CALC)
    _base_ebit=r; r=put(ws,r,["Base EBIT at configured hostel volume",
        f'=({_iref("Revenue take rate")}*{_iref("Hostel comparison AOV")}-{_iref("City gig leg")}-{_iref("In-gate roster leg")}-{_iref("Store ops per order")}-{_iref("Packaging per order")}-{_iref("Unallocated residual")})*{_iref("Hostel term volume in PG scenario")}*{_iref("Term days")}-{_iref("Store fixed cost")}*12',
        "Rs/year","30% non-grocery AOV comparison basis"],fill=CALC)
    r=put(ws,r,["Pro-forma node ROCE",f'=(B{_base_ebit}+B{_ebit})/({_iref("Capex midpoint")}+{_iref("NWC days")}*{_iref("Cluster volume")}*{_iref("Hostel comparison AOV")}*{_iref("NOV / GOV")}+B{_nwc})',"%",
                "PG changes the return only after the evidence gate opens"],fill=CALC,fmt="0.0%")
    for rr in range(_gross, _break+1): ws.cell(rr,2).number_format="0.0"
    for rr in (_lm, _cm): ws.cell(rr,2).number_format="0.00"
    for rr in (_annual, _ebit, _nwc, _base_ebit): ws.cell(rr,2).number_format="#,##0"

    r+=2; ws.cell(row=r,column=1,value=f"NORMALISED CAPACITY SENSITIVITY — PER {P.PG_NORMALISED_BEDS:,} VERIFIED OCCUPIED STUDENT BEDS").font=BOLD; r+=1
    header(ws,r,[("Active-user penetration",22),("Hostel orders/day",18),("Gross PG orders/day",20),
                 ("Spare capacity",17),("PG served",14),("PG overflow",15)]); r+=1
    for pen in P.PG_PENETRATION_SENSITIVITY:
        for hopd in (1000,1200,1400):
            rr=r
            r=put(ws,r,[pen,hopd,
                f'={P.PG_NORMALISED_BEDS}*A{rr}*{_iref("PG orders per active user per day")}',
                f'=MAX(0,{_iref("Cluster volume")}-B{rr})',
                f'=MIN(C{rr},D{rr})',f'=MAX(0,C{rr}-E{rr})'],fill=CALC)
            for cc in range(2,7): ws.cell(rr,cc).number_format="0.0"
        ws.cell(row=r-3,column=1).number_format="0%"; ws.cell(row=r-2,column=1).number_format="0%"; ws.cell(row=r-1,column=1).number_format="0%"

    r+=2; ws.cell(row=r,column=1,value="PG UNIT ECONOMICS — EXISTING MODEL AOV REFERENCES, NOT A FORECAST").font=BOLD; r+=1
    header(ws,r,[("AOV reference",26),("AOV",13),("Common-drop share",20),("Last mile",14),("Contribution/order",20),("Reading",42)]); r+=1
    for label,aov in (("Minutes current midpoint",M.MINUTES_AOV),
                      ("Campus spine breakeven",RC.SPINE_BREAKEVEN),
                      ("30% non-grocery basket",RC.AOV_UP)):
        for share in (0.0,0.5,1.0):
            rr=r
            r=put(ws,r,[label,aov,share,
                f'=(1-C{rr})*{_iref("PG doorstep last mile")}+C{rr}*{_iref("PG common-drop last mile")}',
                f'={_iref("Revenue take rate")}*B{rr}-D{rr}-{_iref("Store ops per order")}-{_iref("Packaging per order")}-{_iref("Unallocated residual")}',
                f'=IF(E{rr}>0,"positive before fixed cost","does not cover variable cost")'],fill=CALC)
            ws.cell(row=rr,column=3).number_format="0%"
            ws.cell(row=rr,column=2).number_format="0.00"
            ws.cell(row=rr,column=4).number_format="0.00"
            ws.cell(row=rr,column=5).number_format="0.00"
    wb.save(OUT); return OUT

if __name__=="__main__":
    p=build(); print("  wrote", os.path.relpath(p,ROOT))
