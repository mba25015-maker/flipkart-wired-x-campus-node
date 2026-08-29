import pandas as pd, campus_model as M, index_model as I
from openpyxl.styles import Font, PatternFill, Alignment
import os as _os
_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))  # repo root, portable across sessions
NAVY="FF0D1F5C"; GOLD="FFFFC220"

assump = pd.DataFrame([
 ["Last mile","Rs/order",M.LAST_MILE,"T2","Industry breakdown of Blinkit economics (7% of order value)"],
 ["Store ops (high)","Rs/order",M.STORE_OPS_HI,"T2","Mid-mile + warehousing (6.5%)"],
 ["Store ops (low)","Rs/order",M.STORE_OPS_LO,"T2","Alt. estimate at 1,400 orders/day"],
 ["Packaging + support","Rs/order",M.PACKAGING,"T2","2% of order value"],
 ["Blinkit contribution/order","Rs",M.BLINKIT_CM,"T1","Eternal Q4 FY26"],
 ["Blinkit net AOV","Rs",M.BLINKIT_AOV,"T1","Eternal Q4 FY26"],
 ["Minutes AOV","Rs",M.MINUTES_AOV,"T2","Inc42 Aug 2026, midpoint 750-800"],
 ["Minutes orders/day/store","orders",M.MINUTES_ORD,"T2","Inc42 Aug 2026, midpoint 1,000-1,100"],
 ["Store fixed cost","Rs/month",750000,"T2","Midpoint of Rs 5-10 L"],
 ["Capex per store","Rs",(M.CAPEX_LO+M.CAPEX_HI)/2,"T2","Midpoint of Rs 2.2-2.5 cr"],
 ["Active academic months","months",M.ACTIVE_MONTHS,"T1","Institutional calendars (8-9)"],
 ["DERIVED blended take rate","%",round(M.TAKE_RATE,4),"D","(Blinkit CM + variable stack) / Blinkit AOV"],
 ["DERIVED calendar surcharge","x",round(M.CAL_SURCHARGE,3),"D","12 / 8.5 active months"],
 ["ASSUMED orders/resident/day","orders",0.25,"A","No source exists. Sensitivity-tested."],
], columns=["Input","Unit","Value","Tier","Source / derivation"])

gate = pd.DataFrame([
 ["Breakeven orders/day (8.5-mo basis)",round(I.OD_BE)],
 ["Orders per resident per day (assumed)",I.ORD_RES],
 ["Minimum viable CLUSTER (hostel residents in one radius)",round(I.CLUSTER)],
 ["Minimum clusters for a state network",I.MIN_CLUSTERS],
 ["State screening threshold (hostel residents)",round(I.GATE)],
 ["States clearing the gate",f"{len(I.g)} of {len(I.df)}"],
], columns=["Step","Value"])

sheets={"1 Assumptions":assump,"2 Required AOV":M.req_aov,"3 Waterfall":M.wf,
        "4 Order density":M.density,"5 Residents required":M.residents,
        "6 Asset turn":M.asset,"7 Take-rate sensitivity":M.sens,
        "8 Scale gate":gate,
        "9 Campus Opportunity Index":I.g[["Rank","State","Concentration","MPCE","Residents",
                                          "Clusters supportable","COI"]],
        "10 Excluded sub-scale":pd.DataFrame({"State":I.df[~I.df["Clears gate"]].State,
                                              "Hostel residents":I.df[~I.df["Clears gate"]].Residents})}
p=_os.path.join(_ROOT, "Campus_Store_Model.xlsx")
with pd.ExcelWriter(p, engine="openpyxl") as w:
    for n,d in sheets.items():
        d.to_excel(w,sheet_name=n[:31],index=False)
        ws=w.sheets[n[:31]]
        for cell in ws[1]:
            cell.font=Font(bold=True,color="FFFFFFFF"); cell.fill=PatternFill("solid",fgColor=NAVY)
            cell.alignment=Alignment(vertical="center",wrap_text=True)
        ws.freeze_panes="A2"
        for col in ws.columns:
            L=max((len(str(c.value)) for c in col if c.value is not None),default=8)
            ws.column_dimensions[col[0].column_letter].width=min(max(L+2,11),52)
print("wrote",p)
